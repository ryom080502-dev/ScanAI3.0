import os
import json
import time
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
import streamlit as st
import google.generativeai as genai
from dotenv import load_dotenv

# --- è¨­å®š ---
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

# â–¼â–¼â–¼ ãƒ¢ãƒ‡ãƒ«æŒ‡å®š â–¼â–¼â–¼
MODEL_NAME = "gemini-3-pro-preview"
TEMPLATE_FILE = "template.xlsx"

# â–¼â–¼â–¼ åˆè¨€è‘‰ã®è¨­å®š â–¼â–¼â–¼
LOGIN_PASSWORD = "fujishima8888" 

# --- ãƒšãƒ¼ã‚¸è¨­å®š ---
st.set_page_config(page_title="çµŒè²»ç²¾ç®—ã‚¹ã‚­ãƒ£ãƒ³AI", layout="wide")

# â–¼â–¼â–¼ CSSã‚¹ã‚¿ã‚¤ãƒ« â–¼â–¼â–¼
st.markdown("""
    <style>
    [data-testid="stFileUploaderDropzoneInstructions"] > div > span {display: none;}
    [data-testid="stFileUploaderDropzoneInstructions"] > div::after { content: "ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ‰ãƒ©ãƒƒã‚°ã¾ãŸã¯é¸æŠ"; font-weight: bold; font-size: 1rem; }
    [data-testid="stFileUploaderDropzoneInstructions"] > div > small {display: none;}
    [data-testid="stFileUploaderDropzoneInstructions"] > div::before { content: "ä¸Šé™ 200MB / PDFã®ã¿"; font-size: 0.8rem; display: block; margin-bottom: 5px; }
    [data-testid="stMetric"] { background-color: #f0f2f6; padding: 15px; border-radius: 10px; border: 1px solid #e0e0e0; }
    @media (prefers-color-scheme: dark) { [data-testid="stMetric"] { background-color: #262730; border: 1px solid #41444e; } }
    </style>
""", unsafe_allow_html=True)

# --- èªè¨¼æ©Ÿèƒ½ ---
def check_password():
    if 'authenticated' not in st.session_state: st.session_state['authenticated'] = False
    if st.session_state['authenticated']: return True
    st.title("ğŸ”’ ãƒ­ã‚°ã‚¤ãƒ³")
    password = st.text_input("ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„", type="password")
    if st.button("ãƒ­ã‚°ã‚¤ãƒ³"):
        if password == LOGIN_PASSWORD:
            st.session_state['authenticated'] = True
            st.rerun()
        else: st.error("ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒé•ã„ã¾ã™")
    return False

# --- çµåˆã‚»ãƒ«å¯¾å¿œæ›¸ãè¾¼ã¿ ---
def smart_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    else: cell.value = value

# --- é›†è¨ˆãƒ»åˆ†é¡ãƒ­ã‚¸ãƒƒã‚¯ ---
def aggregate_receipt_data(raw_data):
    """
    ãƒ‡ãƒ¼ã‚¿ã‚’ã€Œäº¤é€šè²»ã€ã€Œé§è»Šå ´ã€ã€Œé«˜é€Ÿä»£ã€ã€Œä¸€èˆ¬ã€ã®4ã¤ã«åˆ†é¡ã—ã¦é›†è¨ˆã™ã‚‹
    """
    df = pd.DataFrame(raw_data)
    if df.empty: 
        return {"transport": None, "parking": None, "highway": None, "general": []}

    # æ•°å€¤å¤‰æ›
    cols_to_num = ['total_amount', 'amount_8_percent']
    for col in cols_to_num:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    result_dict = {
        "transport": None, # 9è¡Œç›®ç”¨
        "parking": None,   # 10è¡Œç›®ç”¨
        "highway": None,   # 11è¡Œç›®ä»¥é™ã®å…ˆé ­
        "general": []      # 11è¡Œç›®ä»¥é™
    }

    # --- 1. äº¤é€šè²» (transport) ---
    df_trans = df[df['category'] == 'transport']
    if not df_trans.empty:
        total = df_trans['total_amount'].sum()
        total_8 = df_trans['amount_8_percent'].sum()
        latest_date = df_trans['date'].max()
        
        result_dict["transport"] = {
            "date": latest_date,
            "store_name": "äº¤é€šè²»ï¼ˆé›»è»Šãƒ»ãƒã‚¹ç­‰ï¼‰",
            "invoice_number": "", 
            "total_amount": total,
            "amount_8_percent": total_8
        }

    # --- 2. é§è»Šå ´ (parking) ---
    df_park = df[df['category'] == 'parking']
    if not df_park.empty:
        total = df_park['total_amount'].sum()
        total_8 = df_park['amount_8_percent'].sum()
        latest_date = df_park['date'].max()
        
        result_dict["parking"] = {
            "date": latest_date,
            "store_name": "é§è»Šå ´ä»£",
            "invoice_number": "", 
            "total_amount": total,
            "amount_8_percent": total_8
        }

    # --- 3. é«˜é€Ÿä»£ (highway) ---
    df_high = df[df['category'] == 'highway']
    if not df_high.empty:
        total = df_high['total_amount'].sum()
        total_8 = df_high['amount_8_percent'].sum()
        latest_date = df_high['date'].max()
        
        result_dict["highway"] = {
            "date": latest_date,
            "store_name": "é«˜é€Ÿä»£",
            "invoice_number": "", 
            "total_amount": total,
            "amount_8_percent": total_8
        }

    # --- 4. ä¸€èˆ¬ (general) ---
    df_gen = df[~df['category'].isin(['transport', 'parking', 'highway'])]
    
    if not df_gen.empty:
        grouped = df_gen.groupby('store_name').agg({
            'date': 'max',
            'total_amount': 'sum',
            'amount_8_percent': 'sum',
            'invoice_number': 'first'
        }).reset_index()

        general_list = []
        for _, row in grouped.iterrows():
            general_list.append({
                "date": row['date'],
                "store_name": row['store_name'],
                "invoice_number": row['invoice_number'],
                "total_amount": row['total_amount'],
                "amount_8_percent": row['amount_8_percent']
            })
        
        general_list.sort(key=lambda x: x.get("date") if x.get("date") else "9999/99/99")
        result_dict["general"] = general_list

    return result_dict

# --- ãƒ¡ã‚¤ãƒ³ãƒ­ã‚¸ãƒƒã‚¯ ---
def analyze_and_create_excel(uploaded_file, template_path, output_excel_path):
    api_key_to_use = API_KEY or st.secrets.get("GOOGLE_API_KEY")
    if not api_key_to_use:
        st.error("APIã‚­ãƒ¼è¨­å®šã‚¨ãƒ©ãƒ¼")
        return None

    genai.configure(api_key=api_key_to_use)
    
    # â–¼â–¼â–¼ ãƒ—ãƒ­ãƒ³ãƒ—ãƒˆ â–¼â–¼â–¼
    model = genai.GenerativeModel(
        model_name=MODEL_NAME,
        generation_config={"temperature": 0, "response_mime_type": "application/json"},
        system_instruction="""
        ã‚ãªãŸã¯æœ€é«˜ãƒ¬ãƒ™ãƒ«ã®ç²¾åº¦ã‚’æŒã¤çµŒç†æ‹…å½“AIã§ã™ã€‚
        ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã•ã‚ŒãŸPDFï¼ˆè¤‡æ•°æšã®ãƒ¬ã‚·ãƒ¼ãƒˆç”»åƒï¼‰ã‹ã‚‰æƒ…å ±ã‚’æŠ½å‡ºã—ã€JSONãƒ‡ãƒ¼ã‚¿ã‚’ä½œæˆã—ã¦ãã ã•ã„ã€‚
        
        ### 1. åº—èˆ—åã®æ­£è¦åŒ–
        - æ”¯åº—åã¯å‰Šé™¤ã—ã€ä¼šç¤¾åã®ã¿æŠ½å‡ºï¼ˆä¾‹: "å³¶å¿  ã€‡ã€‡åº—" â†’ "å³¶å¿ "ï¼‰ã€‚
        
        ### 2. ã‚«ãƒ†ã‚´ãƒªåˆ¤å®š (å„ªå…ˆé †ä½é †)
        
        **A. å…¬å…±äº¤é€šæ©Ÿé–¢ (transport)**
        - é›»è»Šã€ãƒã‚¹ã€åœ°ä¸‹é‰„ã€ãƒ¢ãƒãƒ¬ãƒ¼ãƒ«ã®ã¿ã€‚
        - ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰: ã€Œä¹—è»Šåˆ¸ã€ã€Œåˆ‡ç¬¦ã€ã€Œé‹è³ƒã€ã€Œãƒãƒ£ãƒ¼ã‚¸ã€ã€ŒSuicaã€ã€ŒPASMOã€ã€ŒJRã€ã€Œé§…ã€ã€Œäº¤é€šå±€ã€ã€Œãƒã‚¹ã€ã€‚
        - (é«˜é€Ÿé“è·¯ã‚„ã‚¿ã‚¯ã‚·ãƒ¼ã¯å«ã‚ãªã„)
        
        **B. é«˜é€Ÿé“è·¯ (highway)**
        - ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰: ã€ŒETCã€ã€Œé«˜é€Ÿã€ã€Œæ–™é‡‘æ‰€ã€ã€Œé€šè¡Œæ–™ã€ã€Œæœ‰æ–™é“è·¯ã€ã€ŒHighwayã€ã€Œé¦–éƒ½é«˜ã€ã€‚
        
        **C. é§è»Šå ´ (parking)**
        - ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰: ã€Œé§è»Šå ´ã€ã€Œãƒ‘ãƒ¼ã‚­ãƒ³ã‚°ã€ã€ŒParkingã€ã€Œï¼°ã€ã€Œã‚³ã‚¤ãƒ³ãƒ‘ãƒ¼ã‚­ãƒ³ã‚°ã€ã€‚
        - æ–‡è„ˆ: ã€Œå…¥åº«ã€ã€Œå‡ºåº«ã€ã€Œé§è»Šæ™‚é–“ã€ã®è¨˜è¼‰ãŒã‚ã‚Œã°åº—åã«é–¢ã‚ã‚‰ãšparkingã¨åˆ¤å®šã€‚
        
        **D. ãã®ä»– (general)**
        - ä¸Šè¨˜ä»¥å¤–ã¯ `general`ã€‚

        ### 3. é‡‘é¡ã¨ã‚¤ãƒ³ãƒœã‚¤ã‚¹
        - **date:** YYYY/MM/DDã€‚
        - **total_amount:** æ”¯æ‰•ç·é¡ï¼ˆç¨è¾¼ï¼‰ã€‚
        - **amount_8_percent:** ã€Œ8%å¯¾è±¡ã€ç­‰ã®è¨˜è¼‰ãŒã‚ã‚‹é‡‘é¡ã€‚ãªã‘ã‚Œã° 0ã€‚
        - **invoice_number:** T+13æ¡ã€‚ãªã‘ã‚Œã° nullã€‚
        
        ### å‡ºåŠ›ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆ
        [{"status": "success", "date": "YYYY/MM/DD", "store_name": "...", "category": "general", "invoice_number": "T...", "total_amount": 1000, "amount_8_percent": 0}]
        """
    )

    try:
        temp_pdf_path = "temp_input.pdf"
        with open(temp_pdf_path, "wb") as f: f.write(uploaded_file.getbuffer())

        sample_file = genai.upload_file(path=temp_pdf_path, display_name="User Upload PDF")
        
        with st.spinner(f'èª­ã¿å–ã‚Šä¸­...'):
            while sample_file.state.name == "PROCESSING":
                time.sleep(1)
                sample_file = genai.get_file(sample_file.name)
            
            if sample_file.state.name == "FAILED": return None

            response = model.generate_content([sample_file, "å…¨ãƒšãƒ¼ã‚¸ã®ãƒ¬ã‚·ãƒ¼ãƒˆæƒ…å ±ã‚’æŠ½å‡ºã—ã¦ãã ã•ã„ã€‚"])
            raw_data = json.loads(response.text)

        analyzed_data = aggregate_receipt_data(raw_data)

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active 
        
        # --- æ›¸ãè¾¼ã¿ç”¨ãƒ˜ãƒ«ãƒ‘ãƒ¼é–¢æ•° ---
        def write_row(row_idx, item_data, skip_basic_info=False):
            """
            skip_basic_info=True ã®å ´åˆã€æ—¥ä»˜ã¨åº—åã®æ›¸ãè¾¼ã¿ã‚’ã‚¹ã‚­ãƒƒãƒ—ã™ã‚‹ï¼ˆé‡‘é¡ã®ã¿æ›¸ãï¼‰
            """
            if not item_data: return
            
            # æ—¥ä»˜ãƒ»åº—åã®æ›¸ãè¾¼ã¿åˆ¶å¾¡
            if not skip_basic_info:
                if item_data.get("date"): smart_write(ws, row_idx, 2, item_data["date"])
                if item_data.get("store_name"): smart_write(ws, row_idx, 5, item_data["store_name"])
            
            # é‡‘é¡ã¯å¸¸ã«æ›¸ãè¾¼ã‚€
            total = item_data.get("total_amount", 0)
            amt_8 = item_data.get("amount_8_percent", 0)
            amt_10_target = total - amt_8

            if amt_8 > 0: smart_write(ws, row_idx, 16, amt_8)
            if amt_10_target > 0: smart_write(ws, row_idx, 19, amt_10_target)

        # â–¼â–¼â–¼ æ›¸ãè¾¼ã¿å‡¦ç† â–¼â–¼â–¼
        
        # 1. å…¬å…±æ©Ÿé–¢ (9è¡Œç›®) -> ä¿®æ­£: skip_basic_info=True ã«è¨­å®š
        if analyzed_data["transport"]:
            write_row(9, analyzed_data["transport"], skip_basic_info=True)
            
        # 2. é§è»Šå ´ (10è¡Œç›®) -> skip_basic_info=True (ç¶­æŒ)
        if analyzed_data["parking"]:
            write_row(10, analyzed_data["parking"], skip_basic_info=True)

        # 3. é«˜é€Ÿä»£ã¨ä¸€èˆ¬ (11è¡Œç›®ä»¥é™)
        items_to_write = []
        if analyzed_data["highway"]:
            items_to_write.append(analyzed_data["highway"])
        items_to_write.extend(analyzed_data["general"])

        current_row = 11
        for item in items_to_write:
            if current_row >= 30 and current_row < 41:
                current_row = 41
            write_row(current_row, item)
            current_row += 1

        wb.save(output_excel_path)
        
        display_list = []
        if analyzed_data["transport"]: display_list.append(analyzed_data["transport"])
        if analyzed_data["parking"]: display_list.append(analyzed_data["parking"])
        if analyzed_data["highway"]: display_list.append(analyzed_data["highway"])
        display_list.extend(analyzed_data["general"])
        
        return display_list

    except Exception as e:
        st.error(f"ã‚·ã‚¹ãƒ†ãƒ ã‚¨ãƒ©ãƒ¼: {e}")
        return None

# --- UIå®Ÿè£… ---
if check_password():
    st.title("ğŸ§¾ çµŒè²»ç²¾ç®—ã‚¹ã‚­ãƒ£ãƒ³AI ")
    st.caption(f"Powered by {MODEL_NAME}")
    st.markdown("---")
    
    col1, col2 = st.columns([1, 2.5])

    with col1:
        st.subheader("ğŸ“‚ ãƒ•ã‚¡ã‚¤ãƒ«é¸æŠ")
        uploaded_file = st.file_uploader("PDFã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰", type=["pdf"])
        if uploaded_file:
            st.success("æº–å‚™å®Œäº†")
            st.markdown("""
            **å‡ºåŠ›ãƒ«ãƒ¼ãƒ«:**
            - **09è¡Œç›®:** äº¤é€šè²» (é‡‘é¡ã®ã¿å‡ºåŠ›ãƒ»æ—¥ä»˜ç­‰ã¯ç¶­æŒ)
            - **10è¡Œç›®:** é§è»Šå ´ä»£ (é‡‘é¡ã®ã¿å‡ºåŠ›ãƒ»æ—¥ä»˜ç­‰ã¯ç¶­æŒ)
            - **11è¡Œç›®:** é«˜é€Ÿä»£ (ã‚ã‚Œã°å…ˆé ­)
            - **11è¡Œç›®~:** ãã®ä»– (åº—èˆ—ã”ã¨)
            """)
            if st.button("èª­ã¿å–ã‚Šé–‹å§‹", type="primary", use_container_width=True):
                if os.path.exists(TEMPLATE_FILE):
                    result = analyze_and_create_excel(uploaded_file, TEMPLATE_FILE, "result_download.xlsx")
                    if result:
                        st.session_state['result_data'] = result
                        st.session_state['excel_ready'] = True
                else: st.error("ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
            
            if 'excel_ready' in st.session_state:
                with open("result_download.xlsx", "rb") as f:
                    st.download_button("ğŸ“¥ Excelãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰", f, file_name="çµŒè²»ç²¾ç®—.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="secondary", use_container_width=True)

    with col2:
        st.subheader("ğŸ“Š è§£æçµæœ")
        if 'result_data' in st.session_state:
            data = st.session_state['result_data']
            total = sum([d.get("total_amount", 0) for d in data])
            st.metric("æ”¯æ‰•ç·é¡", f"Â¥{total:,}")
            
            df = pd.DataFrame(data)
            df["val_10"] = df["total_amount"] - df["amount_8_percent"]
            
            def get_icon(cat_name):
                s = str(cat_name)
                if "äº¤é€šè²»" in s: return "ğŸš†"
                if "é§è»Šå ´" in s: return "ğŸ…¿ï¸"
                if "é«˜é€Ÿä»£" in s: return "ğŸ›£ï¸"
                return "ğŸ›’"

            df["Type"] = df["store_name"].apply(get_icon)
            
            st.dataframe(
                df[["Type", "date", "store_name", "total_amount", "val_10", "amount_8_percent"]].rename(columns={"date":"æ—¥ä»˜","store_name":"é …ç›®/åº—èˆ—å","total_amount":"ç·é¡","val_10":"10%","amount_8_percent":"8%"}),
                use_container_width=True, hide_index=True
            )
        else:
            st.info("å·¦ã®ãƒœã‚¿ãƒ³ã§å®Ÿè¡Œã—ã¦ãã ã•ã„")